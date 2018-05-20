import cognitive_face as CF
from global_variables import personGroupId
import os, urllib
import sqlite3
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter, cell, column_index_from_string
import time

import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

#get current date
currentDate = time.strftime("%d_%m_%y")
wb = load_workbook(filename = "reports.xlsx")
sheet = wb['attendance']

def getDateColumn():
	for i in range(1, sheet.max_column + 1):
		col = get_column_letter(i)
		if sheet['%s%s'% (col,'1')].value == currentDate:
			return col


Key = 'e72e6fd9e8964cdab9b1ba2cc1b14c7b'
CF.Key.set(Key)

BASE_URL = 'https://westcentralus.api.cognitive.microsoft.com/face/v1.0'  # Replace with your regional Base URL
CF.BaseUrl.set(BASE_URL)

connect = connect = sqlite3.connect("Face-DataBase")
c = connect.cursor()

attend = [0 for i in range(200)]

currentDir = os.path.dirname(os.path.abspath(__file__))
directory = os.path.join(currentDir, 'Cropped_faces')
for filename in os.listdir(directory):
	if filename.endswith(".jpg"):
		imgurl = urllib.request.pathname2url(os.path.join(directory, filename))
		res = CF.face.detect(imgurl)
		if len(res) != 1:
			print ("No face detected.")
			continue

		faceIds = []
		for face in res:
			faceIds.append(face['faceId'])
		res = CF.face.identify(faceIds, personGroupId)
		print (filename)
		print (res)
		for face  in res:
			if not face['candidates']:
				print ("Unknown")
			else:
				personId = face['candidates'][0]['personId']
				print(personId)
				c.execute("SELECT * FROM Students WHERE personID = ?", (personId,))
				row = c.fetchone()
				print (row)
				attend[int(row[0])-17000] += 1
				print (row[1] + " recognized")

for row in range(2, sheet.max_row + 1):
	rn = sheet['A%s'% row].value
	if rn is not None:
		rn = rn[-5:]
		if attend[int(rn)-17000] != 0:
			col = getDateColumn()
			sheet['%s%s' % (col, str(row))] = 1

wb.save(filename = "reports.xlsx")

import shutil
folder = './Cropped_faces'
for the_file in os.listdir(folder):
    file_path = os.path.join(folder, the_file)
    try:
        if os.path.isfile(file_path):
            os.unlink(file_path)
        #elif os.path.isdir(file_path): shutil.rmtree(file_path)
    except Exception as e:
        print(e)


#currentDir = os.path.dirname(os.path.abspath(__file__))
#imgurl = urllib.pathname2url(os.path.join(currentDir, "1.jpg"))
#res = CF.face.detect(imgurl)
#faceIds = []
#for face in res:
 #   faceIds.append(face['faceId'])

#res = CF.face.identify(faceIds,personGroupId)
# for face in res:
#     personName = CF.person.get(personGroupId, face['candidates']['personId'])
#     print personName
#print res
